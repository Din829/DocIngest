"""
Docling parser — primary document parsing engine.

Wraps IBM Docling to handle PDF, PPTX, XLSX, HTML, images, and more.
Delegates all format routing to Docling internally (no manual extension matching).

If Docling fails, the pipeline falls back to TextParser (FallbackParser).

Design:
  - Lazy import: Docling is imported on first use (heavy library)
  - OCR config: mapped from our YAML to Docling's PdfPipelineOptions
  - Image extraction: Docling extracts images, we optionally describe them with Vision
  - Metadata: extracted from Docling's document model (title, pages, language, etc.)
"""

from __future__ import annotations

import logging
from io import BytesIO
from pathlib import Path
from typing import Any

from .base import BaseParser, ParseResult, PAGEBREAK_MARKER
from ..config import get_nested

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# OCR engine mapping: our config name → Docling option class
# ---------------------------------------------------------------------------

def _resolve_rapidocr_paths(config: dict[str, Any]) -> dict[str, str] | None:
    """
    Resolve the user-configured RapidOCR model paths.

    Returns:
        - dict with all three (det/cls/rec) populated when all three paths
          are set and the files exist on disk
        - None when no paths are configured (caller falls back to RapidOCR's
          default download location, the historical behaviour)

    Raises:
        ValueError when paths are PARTIALLY configured (one or two set,
        others null/empty) — silent partial config is the kind of bug that
        manifests as cryptic "Permission denied" at runtime weeks later.
        Failing loud at construction time is the right call.
        ValueError when paths are set but the files don't exist on disk.

    Config shape (under parsing.ocr.rapidocr_model_paths):
        det: <absolute path to det .onnx>
        cls: <absolute path to cls .onnx>
        rec: <absolute path to rec .onnx>
    """
    paths_cfg = get_nested(config, "parsing.ocr.rapidocr_model_paths", {}) or {}
    raw = {
        "det": paths_cfg.get("det"),
        "cls": paths_cfg.get("cls"),
        "rec": paths_cfg.get("rec"),
    }
    # Treat empty strings the same as null — both mean "not configured".
    configured = {k: v for k, v in raw.items() if v}

    if not configured:
        return None

    if len(configured) < 3:
        missing = [k for k in ("det", "cls", "rec") if k not in configured]
        raise ValueError(
            f"parsing.ocr.rapidocr_model_paths is partially configured "
            f"(missing: {missing}). All three paths (det/cls/rec) must be "
            f"set together, or all three left null. Partial config would "
            f"silently fall back to RapidOCR's default location and likely "
            f"fail with Permission denied in non-root containers."
        )

    from pathlib import Path as _Path
    missing_files = [
        f"{k}={v}" for k, v in configured.items() if not _Path(v).is_file()
    ]
    if missing_files:
        raise ValueError(
            f"parsing.ocr.rapidocr_model_paths points at file(s) that don't "
            f"exist: {missing_files}. Run `rapidocr download_models` in your "
            f"Dockerfile / build step and verify the resulting .onnx paths."
        )

    return configured


def _build_ocr_options(config: dict[str, Any]):
    """
    Build Docling OCR options from our config.

    Two paths:
      1. If parsing.ocr.rapidocr_model_paths is set (all three), we force
         RapidOCR with explicit model paths regardless of engine="auto" —
         the user's intent is unambiguous (they downloaded models, they
         want RapidOCR to use them). Without this, docling's auto-pick
         might land on EasyOCR/Tesseract and silently ignore the paths.
      2. Otherwise, respect engine= as before. engine="auto" returns None
         (let Docling decide); explicit engines build the matching options.

    Returns None when no special handling is needed (engine="auto" with no
    custom paths). Raises ValueError on misconfigured paths — see
    _resolve_rapidocr_paths for details.
    """
    ocr_cfg = get_nested(config, "parsing.ocr", {})
    engine = ocr_cfg.get("engine", "auto")
    languages = ocr_cfg.get("languages", ["eng"])

    # Custom RapidOCR model paths override engine= because their presence is
    # the strongest signal of intent ("I pre-downloaded models, use them").
    custom_rapidocr = _resolve_rapidocr_paths(config)
    if custom_rapidocr is not None:
        try:
            from docling.datamodel.pipeline_options import RapidOcrOptions
            return RapidOcrOptions(
                det_model_path=custom_rapidocr["det"],
                cls_model_path=custom_rapidocr["cls"],
                rec_model_path=custom_rapidocr["rec"],
            )
        except ImportError as e:
            # rapidocr is a docling extras dep — if missing, fail loud
            # rather than silently dropping the paths the user explicitly
            # configured.
            raise ImportError(
                f"parsing.ocr.rapidocr_model_paths was set but RapidOCR is "
                f"not installed: {e}. Install with: pip install rapidocr"
            ) from e

    if engine == "auto":
        # Let Docling pick the best available OCR engine
        return None

    # Map our engine names to Docling option classes
    try:
        if engine == "easyocr":
            from docling.datamodel.pipeline_options import EasyOcrOptions
            return EasyOcrOptions(lang=languages)
        elif engine == "tesseract":
            from docling.datamodel.pipeline_options import TesseractOcrOptions
            return TesseractOcrOptions(lang=languages)
        elif engine == "rapidocr":
            from docling.datamodel.pipeline_options import RapidOcrOptions
            return RapidOcrOptions()
        else:
            logger.warning(f"Unknown OCR engine '{engine}', using Docling default")
            return None
    except ImportError as e:
        logger.warning(f"OCR engine '{engine}' not available: {e}. Using Docling default.")
        return None


# ---------------------------------------------------------------------------
# Docling Parser
# ---------------------------------------------------------------------------

class DoclingParser(BaseParser):
    """
    Primary parser using IBM Docling.

    Handles 15+ document formats with AI layout analysis and table extraction.
    """

    def __init__(self, config: dict[str, Any]) -> None:
        super().__init__(config)
        self._converter = None  # Lazy init

    def _get_converter(self):
        """Lazy-initialize DocumentConverter (heavy import)."""
        if self._converter is not None:
            return self._converter

        from docling.document_converter import DocumentConverter, PdfFormatOption
        from docling.datamodel.pipeline_options import PdfPipelineOptions
        from docling.datamodel.base_models import InputFormat

        # Build PDF pipeline options from config
        pipeline_options = PdfPipelineOptions()

        # OCR
        ocr_cfg = get_nested(self.config, "parsing.ocr", {})
        # do_ocr is config-driven (default True = backward compatible). Docling's
        # built-in OCR (RapidOCR/EasyOCR) only fills image-only regions; on docs
        # whose image content is already handled by per-page Vision it is
        # redundant, and with a CJK-mismatched default model (RapidOCR ships a
        # Chinese model) it actively corrupts Japanese into Simplified-Chinese
        # garbage. Callers disable it via parsing.ocr.enabled.
        pipeline_options.do_ocr = bool(ocr_cfg.get("enabled", True))
        if ocr_cfg.get("force", False):
            # Force OCR: bypass text layer, useful for broken PDFs
            # Attribute name varies by Docling version
            if hasattr(pipeline_options, "force_full_page_ocr"):
                pipeline_options.force_full_page_ocr = True
            elif hasattr(pipeline_options, "force_backend_text"):
                pipeline_options.force_backend_text = False  # Disable text layer → forces OCR

        ocr_options = _build_ocr_options(self.config)
        if ocr_options is not None:
            pipeline_options.ocr_options = ocr_options

        # Table extraction
        pipeline_options.do_table_structure = get_nested(
            self.config, "parsing.pdf.table_extraction", True
        )

        # Image extraction (generate images for Vision processing later)
        pipeline_options.generate_picture_images = get_nested(
            self.config, "parsing.pdf.image_extraction", True
        )

        # Page-level image generation (for per-page Vision)
        vision_enabled = get_nested(self.config, "parsing.vision.enabled", True)
        if vision_enabled:
            pipeline_options.generate_page_images = True
            # Docling's images_scale is a multiplier of the PDF's native 72 DPI.
            # 180 DPI → scale 2.5 → 1488×2105 for A4, ~3.1 Mpx (under 4MP cap).
            image_dpi = get_nested(self.config, "parsing.vision.image_dpi", 180)
            pipeline_options.images_scale = image_dpi / 72.0

        # Build converter with PDF options
        self._converter = DocumentConverter(
            format_options={
                InputFormat.PDF: PdfFormatOption(
                    pipeline_options=pipeline_options
                ),
            }
        )
        return self._converter

    def parse(
        self,
        file_path: Path,
        *,
        override_stream: BytesIO | None = None,
    ) -> ParseResult:
        """
        Parse a document using Docling.

        Args:
            file_path: Path to the input file (used for naming, metadata,
                and format detection even when override_stream is provided).
            override_stream: Optional BytesIO stream to feed Docling instead
                of reading file_path. Used by pre-parse hooks (e.g. DOCX OMML
                preprocessing) to transform file content before Docling sees
                it without touching the original file on disk.
        """
        from .base import PageData

        # xlsx pre-route: Docling's Excel backend misaligns sheet bodies
        # against their `## SheetName` headers (a sheet's content can end up
        # under the previous sheet's heading), which breaks title_path
        # routing for chunks. Render xlsx directly via openpyxl instead —
        # every sheet's body is guaranteed to live under its own heading.
        # All other formats keep going through Docling unchanged.
        # Config knob (default ON): parsing.xlsx.use_openpyxl_renderer.
        # .xls (legacy BIFF) is handled upstream by pipeline.py's Phase 0.5
        # auto-convert to .xlsx, so by the time we get here the suffix is
        # always .xlsx — no need to test for ".xls" again.
        if (
            file_path.suffix.lower() == ".xlsx"
            and get_nested(self.config, "parsing.xlsx.use_openpyxl_renderer", True)
        ):
            xlsx_result = self._parse_xlsx_via_openpyxl(file_path, override_stream)
            if xlsx_result is not None and xlsx_result.success:
                return xlsx_result
            # Renderer unavailable (openpyxl missing) or refused (file broken)
            # → fall through to the Docling path so we still get *something*
            # rather than failing the file. A debug log is emitted by the
            # helper to make this observable.

        try:
            converter = self._get_converter()
            if override_stream is not None:
                # Route through DocumentStream so Docling reads the
                # transformed bytes while still seeing the original filename
                # (which it uses for format detection).
                from docling_core.types.io import DocumentStream
                override_stream.seek(0)
                doc_stream = DocumentStream(
                    name=file_path.name,
                    stream=override_stream,
                )
                result = converter.convert(doc_stream)
            else:
                result = converter.convert(str(file_path))
            doc = result.document

            # Export full Markdown (with page break markers)
            try:
                markdown = doc.export_to_markdown(
                    page_break_placeholder=PAGEBREAK_MARKER,
                )
            except TypeError:
                markdown = doc.export_to_markdown()

            # Inject section names from Docling groups (sheet names, slide titles, etc.)
            # This makes grep/search find content by real section names
            markdown = self._inject_section_names(doc, markdown)

            # Extract metadata
            metadata = self._extract_metadata(doc, file_path)

            # Build per-page data (text + image path) for Vision enrichment
            self._last_parse_metadata = {}
            pages = self._build_page_data(doc, file_path)

            # Merge xlsx embedded image paths into metadata (if any were extracted)
            if self._last_parse_metadata.get("xlsx_embedded_images"):
                metadata["xlsx_embedded_images"] = self._last_parse_metadata["xlsx_embedded_images"]

            # Extract per-element bounding boxes (if enabled)
            if get_nested(self.config, "output.include_bounding_boxes", True):
                metadata["element_boxes"] = self._extract_bounding_boxes(doc)

            # Detect hidden text via content_layer (if enabled, PDF only)
            if get_nested(self.config, "parsing.pdf.hidden_text_detection.enabled", True):
                hidden_info = self._detect_hidden_content(doc)
                if hidden_info["hidden_element_count"] > 0:
                    metadata["hidden_text"] = hidden_info

            return ParseResult(
                markdown=markdown,
                metadata=metadata,
                pages=pages,
                success=True,
            )
        except Exception as e:
            logger.error(f"Docling parse failed for {file_path.name}: {e}")
            return ParseResult(
                markdown="",
                success=False,
                error=f"Docling parse failed: {e}",
            )

    @staticmethod
    def _extract_bounding_boxes(doc) -> dict:
        """
        Extract per-element bounding boxes from Docling's Document model.

        Returns a dict organized by page number::

            {
                1: [
                    {"label": "text", "bbox": [l, t, r, b], "text_preview": "..."},
                    {"label": "table", "bbox": [l, t, r, b]},
                ],
                2: [...],
            }

        Coordinates are in Docling's coordinate system (origin depends on
        document, typically top-left for PDF). Only elements with provenance
        data are included.
        """
        boxes: dict[int, list[dict]] = {}
        try:
            for item, _level in doc.iterate_items():
                if not item.prov:
                    continue
                prov = item.prov[0]
                page_no = prov.page_no
                b = prov.bbox
                entry: dict = {
                    "label": item.label.value,
                    "bbox": [round(b.l, 1), round(b.t, 1), round(b.r, 1), round(b.b, 1)],
                }
                # Include short text preview for text elements (aids debugging)
                if hasattr(item, "text") and item.text:
                    entry["text_preview"] = item.text[:60]
                boxes.setdefault(page_no, []).append(entry)
        except Exception as e:
            logger.debug(f"Bounding box extraction failed: {e}")
        return boxes

    @staticmethod
    def _detect_hidden_content(doc) -> dict:
        """
        Detect hidden content using Docling's ContentLayer metadata.

        Checks each element's content_layer field — Docling marks some
        elements as INVISIBLE (e.g. Excel hidden sheets) or BACKGROUND
        (watermarks). This is a lightweight check that uses only the
        high-level Document API (no low-level PDF cell access needed).

        For deeper hidden text detection (rendering mode, font color vs
        background), PyMuPDF would be needed — but that's a heavier
        operation reserved for future enhancement.

        Returns::

            {
                "hidden_element_count": int,
                "background_element_count": int,
                "details": [{"page": int, "label": str, "layer": str, "preview": str}],
            }
        """
        from docling_core.types.doc.document import ContentLayer

        hidden_count = 0
        background_count = 0
        details: list[dict] = []
        try:
            for item, _level in doc.iterate_items():
                layer = item.content_layer
                if layer == ContentLayer.INVISIBLE:
                    hidden_count += 1
                    page_no = item.prov[0].page_no if item.prov else 0
                    preview = getattr(item, "text", "")[:60] if hasattr(item, "text") else ""
                    details.append({
                        "page": page_no,
                        "label": item.label.value,
                        "layer": "invisible",
                        "preview": preview,
                    })
                elif layer == ContentLayer.BACKGROUND:
                    background_count += 1
        except Exception as e:
            logger.debug(f"Hidden content detection failed: {e}")

        return {
            "hidden_element_count": hidden_count,
            "background_element_count": background_count,
            "details": details,
        }

    @staticmethod
    def _extract_group_names(doc) -> list[str]:
        """
        Extract section/group names from Docling document structure.

        Works for any format:
          - Excel: "sheet: Day5" → "Day5"
          - PPT: "slide-0" → "Slide 1: <real title>"  (resolved from first text child)
          - PDF: may have chapter names
          - Others: whatever Docling provides

        Internal Docling group names (`slide-0`, `list`, `rich_cell_group_1_0_0`, …)
        carry no semantic value to a downstream RAG / Agent: they would render as
        `## slide-0` headings that bury the real title (which usually appears as a
        bold `#` line inside the section). For groups whose name is a Docling
        internal placeholder, we try to resolve a real title from the first
        meaningful text child; if none is available, we emit an empty name so the
        caller (`_inject_section_names`) drops the heading entirely instead of
        polluting the output with `## list`.

        Detection of "internal placeholder" is purely structural (lowercase
        ASCII + digits + `-_` only) so it works regardless of language and
        does NOT block any real business heading like `Day5`, `Q1 2026`, or
        Japanese / Chinese titles.

        Returns list of names aligned with pagebreak sections. Empty strings
        in the list signal "no usable name — caller should skip injecting a
        heading here".
        """
        try:
            d = doc.export_to_dict()
        except Exception:
            return []

        groups = d.get("groups", [])
        names: list[str] = []

        for g in groups:
            raw_name = g.get("name", "")
            # Sheet names: keep the existing strip.
            if raw_name.startswith("sheet: "):
                names.append(raw_name[7:])
                continue

            # Docling internal placeholder name → try to recover a real title
            # from the group's children. Returns "" if nothing usable found,
            # which tells the caller to skip this section's heading.
            if DoclingParser._looks_like_internal_group_name(raw_name):
                real_title = DoclingParser._first_text_from_children(g, d)
                names.append(real_title)  # may be ""
                continue

            # Real business name (CJK / spaces / mixed case) — keep as-is.
            names.append(raw_name)

        return names

    @staticmethod
    def _looks_like_internal_group_name(name: str) -> bool:
        """
        Heuristic: does this group name look like a Docling-generated
        placeholder rather than a real business label?

        Docling's auto-generated names follow conventions like `slide-0`,
        `list`, `rich_cell_group_1_0_0`, `inline`, `key_value_region`, etc.
        They are ALL lowercase ASCII alphanumerics with `-` / `_` separators.
        A real business label (any sheet/section name a human typed) almost
        always has at least one of: a CJK character, a space, an uppercase
        letter, or punctuation outside `-_`.

        Returns True when we are confident the name is internal junk.
        Returns False (= keep as-is) on any ambiguity — never destroys a
        real name.
        """
        if not name:
            return False  # empty name → caller already handles
        # Real names usually contain a space, CJK, or mixed case → bail out.
        if any(ch.isspace() or ord(ch) > 127 or ch.isupper() for ch in name):
            return False
        # All-lowercase ASCII alphanumeric + - _ → looks generated.
        return all(ch.isalnum() or ch in "-_" for ch in name)

    @staticmethod
    def _first_text_from_children(group: dict, doc_dict: dict, _depth: int = 0) -> str:
        """
        Walk a Docling group's children and return the first meaningful text
        we find, recursing into nested groups when needed.

        Each child is a `{'$ref': '#/texts/N'}` or `{'$ref': '#/groups/M'}`
        reference. We resolve `texts` directly and recurse one level into
        sub-groups (depth-capped to keep this O(N) and crash-proof on
        circular references — Docling shouldn't produce cycles, but
        defensive depth limit costs nothing).

        Returns the stripped text (truncated to a reasonable heading length)
        or "" when no usable text can be found.
        """
        if _depth > 3:
            return ""
        for child in group.get("children", []):
            if not isinstance(child, dict):
                continue
            ref = child.get("$ref") or child.get("cref", "")
            if not isinstance(ref, str) or "/" not in ref:
                continue
            # `#/texts/0` → ('texts', 0)
            parts = ref.lstrip("#/").split("/")
            if len(parts) != 2:
                continue
            collection, raw_idx = parts
            try:
                idx = int(raw_idx)
            except ValueError:
                continue

            if collection == "texts":
                text_node = (doc_dict.get("texts") or [])
                if 0 <= idx < len(text_node):
                    candidate = (text_node[idx].get("text") or "").strip()
                    if candidate:
                        # Heading-length cap: titles rarely exceed ~120 chars;
                        # anything longer is almost certainly body text that
                        # happens to be the first child. Better to truncate
                        # than to dump a paragraph into a `## …` heading.
                        return candidate[:120]
            elif collection == "groups":
                sub_groups = doc_dict.get("groups") or []
                if 0 <= idx < len(sub_groups):
                    nested = DoclingParser._first_text_from_children(
                        sub_groups[idx], doc_dict, _depth=_depth + 1
                    )
                    if nested:
                        return nested
        return ""

    @staticmethod
    def _inject_section_names(doc, markdown: str) -> str:
        """
        Inject group/section names as headings after each pagebreak.

        Before: content1 <!-- pagebreak --> content2
        After:  ## ドリルの実施方法\n\ncontent1 <!-- pagebreak -->\n\n## Day5\n\ncontent2

        Only injects if:
          1. Document has groups with meaningful names
          2. Markdown has pagebreaks matching group count

        Skip-injection guard: when the group name we'd inject is byte-identical
        (after strip) to a heading Docling already wrote at the top of the
        section, we omit our injection to avoid `## Title` immediately followed
        by `# Title`. We compare to the first stripped non-blank line of the
        section and require an EXACT match (modulo leading `#` markers and
        surrounding whitespace) — no fuzzy match, so a real distinct heading
        like `## Agenda Items` next to `# Agenda` is preserved.
        """
        names = DoclingParser._extract_group_names(doc)
        if not names:
            return markdown

        if PAGEBREAK_MARKER not in markdown:
            # No pagebreaks — inject heading at top if there's a name AND
            # it's not already the first heading line of the markdown.
            if names and names[0]:
                if DoclingParser._section_already_has_heading(markdown, names[0]):
                    return markdown
                return f"## {names[0]}\n\n{markdown}"
            return markdown

        sections = markdown.split(PAGEBREAK_MARKER)

        # Names should align with sections (1 name per section)
        # Section count may differ from name count (e.g., empty sections skipped)
        for i in range(min(len(sections), len(names))):
            name = names[i].strip()
            if not name:
                continue
            section = sections[i].strip()
            # Skip injection when Docling already emitted this exact heading
            # at the section's top — happens when the group's first text
            # child IS the slide / section title (we resolved a real title
            # from children, but Docling also renders that same text as a
            # body `#`/`##` line right below the pagebreak).
            if section and DoclingParser._section_already_has_heading(section, name):
                continue
            if section:
                sections[i] = f"\n## {name}\n\n{section}\n"
            else:
                sections[i] = f"\n## {name}\n"

        return PAGEBREAK_MARKER.join(sections)

    @staticmethod
    def _section_already_has_heading(section: str, name: str) -> bool:
        """
        True iff the first non-blank line of `section` is a markdown heading
        whose visible text exactly equals `name` (both stripped).

        Strict exact-match by design — `## X` next to a real different
        heading `# Y` must still get our injection. Only kills the obvious
        `## Title` / `# Title` duplication case where the same string
        appears verbatim twice in a row.
        """
        target = name.strip()
        if not target:
            return False
        for raw in section.split("\n"):
            line = raw.strip()
            if not line:
                continue
            # Only the FIRST non-blank line matters — once we hit non-heading
            # content, the section clearly doesn't start with this heading.
            if not line.startswith("#"):
                return False
            # Strip leading `#`s and one optional space.
            stripped = line.lstrip("#").lstrip()
            return stripped == target
        return False

    def _extract_metadata(self, doc, file_path: Path) -> dict[str, Any]:
        """Extract document metadata from Docling result."""
        # Suffix-based format is the baseline. For files with weak or
        # missing extensions (common after zip expansion, or for
        # renamed/downloaded files), the format detector can override
        # this via content-based identification (magika). Strong
        # extensions are trusted by default — see format_detector for
        # the decision logic.
        suffix_format = file_path.suffix.lstrip(".").lower()
        corrected_format: str | None = None
        try:
            from ..utils.format_detector import detect_format
            corrected_format = detect_format(file_path, self.config)
        except Exception as e:
            logger.debug(f"format_detector failed for {file_path.name}: {e}")

        metadata: dict[str, Any] = {
            "format": corrected_format or suffix_format,
            "title": file_path.stem,
        }
        if corrected_format and corrected_format != suffix_format:
            # Preserve the original suffix so debugging and the quality
            # report can see what magika overrode.
            metadata["suffix_format"] = suffix_format

        try:
            if hasattr(doc, "pages") and doc.pages:
                metadata["pages"] = len(doc.pages)
        except Exception:
            pass

        try:
            md_content = doc.export_to_markdown()
            metadata["has_tables"] = "|" in md_content and "---" in md_content
            metadata["has_images"] = "<!-- image" in md_content
        except Exception:
            pass

        # Surface Docling's own minimal metadata (DoclingDocument.origin +
        # name) so downstream hooks and the frontmatter writer can consume
        # them without re-parsing the file. Docling itself does NOT expose
        # author / creation date — those come from the exiftool hook.
        try:
            origin = getattr(doc, "origin", None)
            if origin is not None:
                docling_origin: dict[str, Any] = {}
                for field in ("filename", "mimetype", "binary_hash", "uri"):
                    value = getattr(origin, field, None)
                    if value:
                        docling_origin[field] = value
                if docling_origin:
                    metadata["docling_origin"] = docling_origin
            doc_name = getattr(doc, "name", None)
            if doc_name:
                metadata["docling_name"] = doc_name
        except Exception:
            pass

        return metadata

    def _build_page_data(self, doc, file_path: Path) -> list:
        """
        Build per-page data: extract text and save page image for each page.

        This gives the pipeline everything it needs for per-page Vision decisions.
        No filtering here — the pipeline/Vision prompt handles all logic.
        """
        from .base import PageData

        pages_data: list = []
        if not hasattr(doc, "pages") or not doc.pages:
            return pages_data

        assets_dir = Path(get_nested(
            self.config, "output.dir", "./knowledge"
        )) / get_nested(self.config, "output.assets_dir", "assets")
        assets_dir.mkdir(parents=True, exist_ok=True)

        # Per-page picture count — an OCR-independent "this page has visual
        # content" signal for triage (see PageData.num_pictures and pipeline's
        # _should_skip_vision). Two complementary sources, OR'd per page:
        #   1. Docling picture elements (works for any format Docling parses).
        #   2. (PDF only) fitz embedded-image count — catches images Docling's
        #      layout pass does NOT register as pictures (figures that used to
        #      surface only via OCR garble). Verified necessary: Docling missed
        #      several image pages that fitz sees.
        pic_per_page: dict[int, int] = {}
        for _pic in getattr(doc, "pictures", None) or []:
            _prov = getattr(_pic, "prov", None)
            if _prov:
                _pn = _prov[0].page_no
                pic_per_page[_pn] = pic_per_page.get(_pn, 0) + 1
        if file_path.suffix.lower() == ".pdf":
            # file_path is always the real PDF here — no pre-parse hook rewrites
            # PDFs (the OMML hook targets DOCX), so reading it directly matches
            # what Docling parsed.
            try:
                import fitz
                _fdoc = fitz.open(file_path)
                for _i, _pg in enumerate(_fdoc, start=1):
                    _n = len(_pg.get_images())
                    if _n:
                        pic_per_page[_i] = max(pic_per_page.get(_i, 0), _n)
                _fdoc.close()
            except Exception as e:
                logger.debug(f"fitz per-page image count failed: {e}")

        for page_no, page in doc.pages.items():
            # Extract per-page text via Docling
            page_text = ""
            try:
                page_text = doc.export_to_markdown(page_no=page_no)
            except Exception:
                pass

            # Save page image (if available)
            image_path = ""
            if page.image is not None:
                pil_img = None
                if hasattr(page.image, "pil_image") and page.image.pil_image:
                    pil_img = page.image.pil_image
                elif hasattr(page.image, "save"):
                    pil_img = page.image

                if pil_img is not None:
                    asset_name = f"{file_path.stem}-page-{page_no:03d}.png"
                    output_path = assets_dir / asset_name
                    try:
                        pil_img.save(str(output_path))
                        image_path = str(output_path)
                    except Exception as e:
                        logger.debug(f"Could not save page image: {e}")

            pages_data.append(PageData(
                page_no=page_no,
                text=page_text,
                image_path=image_path,
                num_pictures=pic_per_page.get(page_no, 0),
            ))

        # Extract embedded images from xlsx zip (xl/media/)
        xlsx_denoise = get_nested(self.config, "parsing.xlsx.denoising", {})
        # .xls is converted to .xlsx upstream (pipeline.py Phase 0.5).
        if xlsx_denoise.get("extract_images", True) and file_path.suffix.lower() == ".xlsx":
            embedded = self._extract_xlsx_images(file_path, assets_dir)
            if embedded:
                parse_meta = getattr(self, "_last_parse_metadata", {})
                parse_meta["xlsx_embedded_images"] = embedded
                self._last_parse_metadata = parse_meta

        # If no page images from Docling (e.g., PPT SimplePipeline),
        # try external conversion as fallback
        has_any_image = any(p.image_path for p in pages_data)
        if not has_any_image and pages_data:
            self._try_external_page_images(file_path, assets_dir, pages_data)

        return pages_data

    @staticmethod
    def _extract_xlsx_images(file_path: Path, assets_dir: Path) -> list[str]:
        """
        Extract embedded images from an xlsx file's zip structure.

        xlsx is a zip containing xl/media/image1.png, image2.jpeg, etc.
        These are the images pasted into cells (screenshots, diagrams).
        Works for ALL xlsx — data-heavy ones simply have no images (returns []).
        """
        import zipfile

        if not zipfile.is_zipfile(str(file_path)):
            return []

        image_exts = (".png", ".jpg", ".jpeg", ".gif", ".bmp", ".tiff", ".webp", ".emf", ".wmf")
        extracted: list[str] = []

        try:
            with zipfile.ZipFile(str(file_path), "r") as zf:
                for name in zf.namelist():
                    if not name.startswith("xl/media/"):
                        continue
                    lower = name.lower()
                    if not any(lower.endswith(ext) for ext in image_exts):
                        continue
                    data = zf.read(name)
                    out_name = f"{file_path.stem}-{Path(name).name}"
                    out_path = assets_dir / out_name
                    out_path.write_bytes(data)
                    extracted.append(str(out_path))
        except Exception as e:
            logger.debug(f"xlsx image extraction failed: {e}")

        if extracted:
            logger.info(f"Extracted {len(extracted)} embedded images from {file_path.name}")
        return extracted

    # -----------------------------------------------------------------
    # xlsx renderer (openpyxl-based)
    # -----------------------------------------------------------------

    def _parse_xlsx_via_openpyxl(
        self,
        file_path: Path,
        override_stream: BytesIO | None,
    ) -> ParseResult | None:
        """
        Render xlsx → markdown using openpyxl, bypassing Docling.

        Why this exists: Docling's Excel backend renders all sheets into one
        flat document and *misaligns* sheet headings against bodies — content
        from sheet N can end up under sheet N-1's `## name` heading. That
        leaks into chunk metadata (`title_path` wrong) and breaks downstream
        per-sheet retrieval. openpyxl reads each sheet's cells directly, so
        every sheet's body is guaranteed to live under its own heading.

        Output contract — identical to the Docling path:
          * markdown: one ``## <sheet name>`` heading per visible sheet,
            sheets separated by PAGEBREAK_MARKER, body rendered as a
            standard Markdown table.
          * metadata: format / title / pages (= sheet count) / has_tables /
            docling_origin (synthesized: mimetype + binary_hash so the
            file_metadata hook and frontmatter writer behave identically).
          * pages: empty list — LibreOffice page-image fallback in the
            pipeline later populates this for Vision enrichment, exactly
            the same as it does for Docling-rendered xlsx today.

        Returns:
          * ParseResult(success=True, ...) on successful render
          * ParseResult(success=False, error=...) when the file refuses to
            open (corrupt / not a real xlsx) — caller will fall back to
            Docling
          * None when openpyxl itself is not installed — caller falls back
            silently (no error surfaced to the user)
        """
        try:
            import openpyxl
        except ImportError:
            logger.debug(
                "openpyxl not installed — falling back to Docling for xlsx. "
                "Install with: pip install openpyxl"
            )
            return None

        try:
            # override_stream wins (lets a pre-parse hook transform the file
            # before we see it, mirroring the Docling path's contract).
            if override_stream is not None:
                override_stream.seek(0)
                wb = openpyxl.load_workbook(
                    override_stream, data_only=True, read_only=False
                )
            else:
                wb = openpyxl.load_workbook(
                    str(file_path), data_only=True, read_only=False
                )

            # Pre-extract embedded images and read anchor info directly
            # from the xlsx OOXML structure. We deliberately *do not* use
            # openpyxl's ``ws._images`` here — it silently drops EMF/WMF
            # at load time AND rewrites surviving entries' ``path`` to a
            # value that no longer matches the real zip media (observed
            # on real spec sheets, see ``_collect_xlsx_image_anchors``
            # docstring). Reading the xml ourselves recovers every
            # embedded picture regardless of format.
            assets_dir = Path(get_nested(
                self.config, "output.dir", "./knowledge"
            )) / get_nested(self.config, "output.assets_dir", "assets")
            assets_dir.mkdir(parents=True, exist_ok=True)

            xlsx_denoise = get_nested(self.config, "parsing.xlsx.denoising", {})
            extracted: list[str] = []
            if xlsx_denoise.get("extract_images", True):
                extracted = self._extract_xlsx_images(file_path, assets_dir)

            # Map zip basename → asset filename written to assets/.
            # ``_extract_xlsx_images`` writes ``{stem}-{zip_basename}``,
            # so stripping the stem prefix gives us the key the OOXML
            # drawing rels uses (e.g. ``image1.emf``).
            stem_prefix = f"{file_path.stem}-"
            asset_by_zip_name: dict[str, str] = {}
            for asset_path in extracted:
                fname = Path(asset_path).name
                if fname.startswith(stem_prefix):
                    asset_by_zip_name[fname[len(stem_prefix):]] = fname
                else:
                    asset_by_zip_name[fname] = fname

            # Read sheet→anchor index from the xlsx OOXML. Best-effort:
            # any failure returns an empty dict, which makes the renderer
            # behave as if there were no images (all extracted media end
            # up in the orphan footer below, so info is never lost).
            anchors_by_sheet = _collect_xlsx_image_anchors(file_path)

            sheet_sections: list[str] = []
            anchored_assets: set[str] = set()
            # Visible sheet names in workbook order. Emitted into metadata
            # so the pipeline's sheet→page mapping (utils.xlsx_sheet_map)
            # can match PDF outline bookmarks back to openpyxl sheet names.
            visible_sheet_names: list[str] = []

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                # Skip hidden sheets so chunked output mirrors what users
                # see when they open the file. (Docling's behaviour for
                # hidden sheets varies; explicit skip keeps us predictable.)
                if getattr(ws, "sheet_state", "visible") != "visible":
                    continue
                visible_sheet_names.append(sheet_name)

                row_to_assets: dict[int, list[str]] = {}
                for anchor_info in anchors_by_sheet.get(sheet_name, []):
                    media = anchor_info.get("media")
                    if not media:
                        continue
                    asset_fname = asset_by_zip_name.get(media)
                    if not asset_fname:
                        # Image declared in OOXML but never extracted —
                        # likely because parsing.xlsx.denoising.extract_images
                        # was disabled. Skip silently; the renderer just
                        # won't emit a marker for it.
                        continue
                    row_1 = anchor_info.get("row")
                    if not isinstance(row_1, int):
                        continue
                    row_to_assets.setdefault(row_1, []).append(asset_fname)
                    anchored_assets.add(asset_fname)

                body_lines = _render_xlsx_sheet_to_markdown(
                    ws,
                    image_anchors=row_to_assets,
                    # Per-sheet orphans are emitted by the workbook-level
                    # footer instead (further down) — passing None here
                    # avoids listing the same orphan on every sheet.
                    orphan_image_names=None,
                )
                if body_lines:
                    sheet_sections.append(
                        f"## {sheet_name}\n\n" + "\n".join(body_lines)
                    )
                else:
                    # Render visible-but-empty sheets as a stub so the
                    # `## name` heading is still present — preserves the
                    # invariant "every sheet has its own section".
                    sheet_sections.append(f"## {sheet_name}\n\n*(empty)*")
            wb.close()

            if not sheet_sections:
                # No visible sheets — let Docling try (it might see hidden ones).
                return ParseResult(
                    markdown="",
                    success=False,
                    error="openpyxl: no visible sheets",
                )

            # Workbook-level orphan footer — images that exist in the xlsx
            # zip but weren't anchored to any visible sheet's row by
            # openpyxl. Common cause: EMF/WMF formats dropped at load.
            # Listed in the LAST sheet section so they live inside a
            # pagebreak segment (consistent with sheet-scoped content) and
            # downstream chunkers / Vision triage can pick them up.
            orphan_files = sorted(
                Path(p).name for p in extracted
                if Path(p).name not in anchored_assets
            )
            if orphan_files and sheet_sections:
                footer_lines = [
                    "",
                    "*Embedded images without resolvable anchor "
                    "(present in the workbook but their cell position "
                    "could not be read — most often EMF/WMF; see "
                    "`assets/`):*",
                ]
                for fname in orphan_files:
                    footer_lines.append(f"- <!-- image: {fname} -->")
                sheet_sections[-1] = (
                    sheet_sections[-1] + "\n" + "\n".join(footer_lines)
                )

            # Sheets joined by PAGEBREAK_MARKER, matching the convention of
            # Docling.export_to_markdown(page_break_placeholder=PAGEBREAK_MARKER).
            markdown = f"\n{PAGEBREAK_MARKER}\n".join(sheet_sections)

            # Build metadata — match Docling-path shape so downstream hooks
            # (file_metadata, frontmatter writer, chunk lineage) need zero
            # special-casing.
            import mimetypes
            import hashlib

            mimetype = mimetypes.guess_type(file_path.name)[0] or \
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            try:
                # Stable 64-bit signature of file contents. Not required to
                # match Docling's own hash byte-for-byte — it's a provenance
                # field, not a cache key. Using SHA-256 truncated to 8 bytes
                # gives us a deterministic int that fits in JSON safely.
                with open(file_path, "rb") as fh:
                    digest = hashlib.sha256(fh.read()).digest()
                binary_hash = int.from_bytes(digest[:8], "big")
            except Exception:
                binary_hash = None

            docling_origin: dict[str, Any] = {
                "filename": file_path.name,
                "mimetype": mimetype,
            }
            if binary_hash is not None:
                docling_origin["binary_hash"] = binary_hash

            metadata: dict[str, Any] = {
                "format": "xlsx",
                "title": file_path.stem,
                "pages": len(sheet_sections),
                "has_tables": True,
                # has_images is true when ANY image marker ends up in the
                # rendered markdown — that's the canonical signal the
                # rest of the pipeline (Vision triage, chunk metadata,
                # quality report) keys off.
                "has_images": "<!-- image" in markdown,
                "docling_origin": docling_origin,
                "docling_name": file_path.stem,
                # Visible sheet count — used by _enrich_with_vision's batched
                # multi-image trigger (pages/sheet ratio). Independent from
                # `pages` above (which is sheet_sections length) so future
                # changes to either don't accidentally break the trigger.
                "xlsx_visible_sheet_count": len(sheet_sections),
                # Visible sheet names in workbook order — fed to
                # utils.xlsx_sheet_map.build_sheet_page_map after the
                # LibreOffice PDF render, so the resulting sheet→page map
                # uses the same names downstream code reads from openpyxl.
                "xlsx_visible_sheet_names": list(visible_sheet_names),
            }
            if extracted:
                metadata["xlsx_embedded_images"] = extracted

            logger.info(
                f"xlsx rendered via openpyxl: {file_path.name} "
                f"→ {len(sheet_sections)} sheet(s), {len(markdown):,} chars"
            )

            return ParseResult(
                markdown=markdown,
                metadata=metadata,
                pages=[],   # LibreOffice fallback in pipeline fills this
                success=True,
            )

        except Exception as e:
            # File broken / not a real xlsx / openpyxl bug — surface for
            # fallback. Don't crash the file; the caller will try Docling.
            logger.warning(
                f"openpyxl xlsx render failed for {file_path.name}: {e}. "
                f"Falling back to Docling."
            )
            return ParseResult(
                markdown="",
                success=False,
                error=f"openpyxl render failed: {e}",
            )

    def _try_external_page_images(
        self, file_path: Path, assets_dir: Path, pages_data: list
    ) -> None:
        """
        Fallback: convert document to page images using external tools.

        Tries LibreOffice headless → PDF → images pipeline.
        Gracefully does nothing if tools aren't available.
        """
        import subprocess
        import tempfile
        from ..utils.binary_finder import find_binary

        # Cross-platform LibreOffice lookup — handles Windows Program Files
        # installs, macOS /Applications bundles, and config/env overrides.
        soffice = find_binary("soffice", self.config)
        if not soffice:
            logger.debug("LibreOffice not found — skipping PPT page image export")
            return

        try:
            with tempfile.TemporaryDirectory() as tmpdir:
                # Step 1: Convert to PDF via LibreOffice
                subprocess.run(
                    [soffice, "--headless", "--convert-to", "pdf",
                     "--outdir", tmpdir, str(file_path)],
                    capture_output=True, timeout=120,
                )
                pdf_files = list(Path(tmpdir).glob("*.pdf"))
                if not pdf_files:
                    return

                # Resolve target DPI from config (unified across all paths)
                image_dpi = get_nested(self.config, "parsing.vision.image_dpi", 180)

                # Step 2: PDF pages → images via Docling or pdf2image
                try:
                    from pdf2image import convert_from_path
                    images = convert_from_path(str(pdf_files[0]), dpi=image_dpi)
                    for i, img in enumerate(images):
                        if i < len(pages_data):
                            name = f"{file_path.stem}-page-{pages_data[i].page_no:03d}.png"
                            out = assets_dir / name
                            img.save(str(out))
                            pages_data[i].image_path = str(out)
                except ImportError:
                    # pdf2image not available — try poppler-free approach
                    # Re-parse the PDF with Docling to get page images
                    from docling.document_converter import DocumentConverter, PdfFormatOption
                    from docling.datamodel.pipeline_options import PdfPipelineOptions
                    from docling.datamodel.base_models import InputFormat

                    opts = PdfPipelineOptions()
                    opts.generate_page_images = True
                    opts.images_scale = image_dpi / 72.0
                    conv = DocumentConverter(
                        format_options={InputFormat.PDF: PdfFormatOption(pipeline_options=opts)}
                    )
                    result = conv.convert(str(pdf_files[0]))
                    for page_no, page in result.document.pages.items():
                        if page.image is None:
                            continue
                        pil = getattr(page.image, "pil_image", None)
                        if pil is None:
                            continue
                        # Match page_no to pages_data index
                        idx = page_no - 1  # Docling pages are 1-indexed
                        if 0 <= idx < len(pages_data):
                            name = f"{file_path.stem}-page-{pages_data[idx].page_no:03d}.png"
                            out = assets_dir / name
                            pil.save(str(out))
                            pages_data[idx].image_path = str(out)

        except Exception as e:
            logger.debug(f"External page image conversion failed: {e}")

    def supported_extensions(self) -> set[str]:
        return {
            ".pdf", ".docx", ".pptx", ".xlsx",
            ".html", ".htm",
            ".png", ".jpg", ".jpeg", ".tiff", ".bmp", ".webp", ".gif",
            ".md", ".txt", ".csv",
            ".asciidoc", ".adoc",
        }


# ---------------------------------------------------------------------------
# xlsx rendering helpers (module-level, pure functions — easy to unit-test)
# ---------------------------------------------------------------------------

# XML namespaces used by the OOXML xlsx format for sheet drawings.
# Defined once here so both the anchor collector and any future helper
# share the same constants without re-declaring them.
_OOXML_NS = {
    "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
}


def _ooxml_normalize_target(target: str, base_path: str) -> str:
    """
    Normalise an OOXML Relationship Target into an absolute zip path.

    Relationships use paths that are either absolute (``/xl/...``) or
    relative to the directory of the referencing xml file. The xlsx zip
    itself stores everything as flat absolute names (without a leading
    slash), so we resolve here:

      ``/xl/worksheets/sheet1.xml``           → ``xl/worksheets/sheet1.xml``
      ``../drawings/drawing1.xml`` from sheet → ``xl/drawings/drawing1.xml``

    Returns the normalised zip-internal path; the caller is responsible
    for checking it actually exists in the zip namelist.
    """
    if target.startswith("/"):
        return target.lstrip("/")
    base_dir = base_path.rsplit("/", 1)[0] if "/" in base_path else ""
    segments = (base_dir.split("/") if base_dir else []) + target.split("/")
    stack: list[str] = []
    for seg in segments:
        if seg in ("", "."):
            continue
        if seg == "..":
            if stack:
                stack.pop()
        else:
            stack.append(seg)
    return "/".join(stack)


def _collect_xlsx_image_anchors(file_path: Path) -> dict[str, list[dict[str, Any]]]:
    """
    Read sheet→image anchor info directly from the xlsx OOXML structure.

    Why this exists (we tried openpyxl first):
      ``openpyxl.load_workbook(...).worksheets[i]._images`` silently drops
      EMF/WMF images at load time ("wmf image format is not supported")
      and additionally rewrites the surviving entries' ``path`` attribute
      to a value that no longer matches the actual media file in the zip
      (observed on real spec sheets: 5 anchors all claiming
      ``/xl/media/image1.png`` when the file is really ``image1.emf``).
      Reading the OOXML directly sidesteps both problems — every image
      anchor declared in ``xl/drawings/drawing*.xml`` becomes available,
      regardless of format.

    What gets collected:
      Only ``<xdr:pic>`` anchors (embedded pictures). Shape anchors
      (``<xdr:sp>``, ``<xdr:cxnSp>`` — text boxes, arrows, connectors
      used in sequence diagrams) are deliberately skipped: they carry no
      media reference and the LibreOffice page-image render already
      describes them via the Vision path. Including them here would
      produce ghost image markers with no corresponding ``assets/`` file.

    Returns:
      {sheet_display_name: [{"row": int, "col": int, "media": str}, ...]}

      ``row`` / ``col`` are 1-indexed (matching ``ws.cell(row=...)``).
      ``media`` is the basename of the embedded media file (e.g.
      ``"image1.emf"``), matching the names emitted by
      ``_extract_xlsx_images``.

      Sheets with no drawing reference get an empty list. Workbooks that
      are not real xlsx (or that fail to open) return ``{}``.

    Errors are swallowed: this is a best-effort enrichment, never a
    correctness gate. If we can't read the anchors, downstream behaviour
    is "no markers inserted, all extracted media listed as orphans"
    — degraded but never broken.
    """
    import xml.etree.ElementTree as ET
    import zipfile

    result: dict[str, list[dict[str, Any]]] = {}
    if not zipfile.is_zipfile(str(file_path)):
        return result

    try:
        zf = zipfile.ZipFile(str(file_path))
    except Exception as exc:
        logger.debug(f"xlsx anchor collect: cannot open zip ({exc})")
        return result

    try:
        names = set(zf.namelist())
        try:
            wb_xml = ET.fromstring(zf.read("xl/workbook.xml"))
            wb_rels_xml = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
        except (KeyError, ET.ParseError) as exc:
            logger.debug(f"xlsx anchor collect: workbook.xml/rels missing or unparseable ({exc})")
            return result

        wb_rels: dict[str, str] = {}
        for rel in wb_rels_xml:
            rid = rel.attrib.get("Id")
            target = rel.attrib.get("Target", "")
            if rid:
                wb_rels[rid] = target

        sheets_elem = wb_xml.find("main:sheets", _OOXML_NS)
        if sheets_elem is None:
            return result

        r_ns = _OOXML_NS["r"]
        a_ns = _OOXML_NS["a"]

        for sh in sheets_elem.findall("main:sheet", _OOXML_NS):
            sheet_name = sh.attrib.get("name", "")
            rid = sh.attrib.get(f"{{{r_ns}}}id")
            if not sheet_name or not rid:
                continue
            sheet_target = wb_rels.get(rid)
            if not sheet_target:
                continue
            sheet_path = _ooxml_normalize_target(sheet_target, "xl/workbook.xml")
            if sheet_path not in names:
                continue

            try:
                sheet_xml = ET.fromstring(zf.read(sheet_path))
            except (KeyError, ET.ParseError):
                continue
            drawing_elem = sheet_xml.find("main:drawing", _OOXML_NS)
            if drawing_elem is None:
                result[sheet_name] = []
                continue
            drawing_rid = drawing_elem.attrib.get(f"{{{r_ns}}}id")
            if not drawing_rid:
                result[sheet_name] = []
                continue

            # Sheet-level rels: rId → drawing xml path
            sheet_dir, _, sheet_fname = sheet_path.rpartition("/")
            sheet_rels_path = f"{sheet_dir}/_rels/{sheet_fname}.rels"
            if sheet_rels_path not in names:
                result[sheet_name] = []
                continue
            try:
                sheet_rels_xml = ET.fromstring(zf.read(sheet_rels_path))
            except (KeyError, ET.ParseError):
                continue
            sheet_rels: dict[str, str] = {}
            for rel in sheet_rels_xml:
                rid_ = rel.attrib.get("Id")
                target = rel.attrib.get("Target", "")
                if rid_:
                    sheet_rels[rid_] = target

            drawing_target = sheet_rels.get(drawing_rid)
            if not drawing_target:
                result[sheet_name] = []
                continue
            drawing_path = _ooxml_normalize_target(drawing_target, sheet_path)
            if drawing_path not in names:
                continue

            # Drawing-level rels: rId → media basename
            drawing_dir, _, drawing_fname = drawing_path.rpartition("/")
            drawing_rels_path = f"{drawing_dir}/_rels/{drawing_fname}.rels"
            drawing_rels: dict[str, str] = {}
            if drawing_rels_path in names:
                try:
                    drx = ET.fromstring(zf.read(drawing_rels_path))
                    for rel in drx:
                        drx_id = rel.attrib.get("Id")
                        drx_target = rel.attrib.get("Target", "")
                        if drx_id and drx_target:
                            # Only need the basename — caller pairs this
                            # with `_extract_xlsx_images`'s output, which
                            # uses ``{stem}-{basename}`` naming.
                            drawing_rels[drx_id] = Path(drx_target).name
                except (KeyError, ET.ParseError):
                    pass

            try:
                drawing_xml = ET.fromstring(zf.read(drawing_path))
            except (KeyError, ET.ParseError):
                continue

            anchors: list[dict[str, Any]] = []
            for anchor_tag in ("twoCellAnchor", "oneCellAnchor"):
                for anc in drawing_xml.findall(f"xdr:{anchor_tag}", _OOXML_NS):
                    from_elem = anc.find("xdr:from", _OOXML_NS)
                    if from_elem is None:
                        continue
                    row_elem = from_elem.find("xdr:row", _OOXML_NS)
                    col_elem = from_elem.find("xdr:col", _OOXML_NS)
                    if row_elem is None or col_elem is None:
                        continue
                    try:
                        row_1 = int((row_elem.text or "0").strip()) + 1
                        col_1 = int((col_elem.text or "0").strip()) + 1
                    except ValueError:
                        continue

                    # Picture anchors only — skip shape (``sp``) and
                    # connector (``cxnSp``) anchors; they have no media
                    # reference and would otherwise produce ghost markers.
                    pic = anc.find("xdr:pic", _OOXML_NS)
                    if pic is None:
                        continue
                    embed_rid = None
                    for blip in pic.iter(f"{{{a_ns}}}blip"):
                        embed_rid = blip.attrib.get(f"{{{r_ns}}}embed")
                        if embed_rid:
                            break
                    if not embed_rid:
                        continue
                    media_basename = drawing_rels.get(embed_rid)
                    if not media_basename:
                        continue
                    anchors.append({
                        "row": row_1,
                        "col": col_1,
                        "media": media_basename,
                    })

            result[sheet_name] = anchors
    finally:
        zf.close()

    return result


def _render_xlsx_sheet_to_markdown(
    ws,
    image_anchors: dict[int, list[str]] | None = None,
    orphan_image_names: list[str] | None = None,
) -> list[str]:
    """
    Render a single openpyxl Worksheet to Markdown table lines.

    Design choices (each defends against a specific failure mode seen on
    real Japanese spec sheets — see ARCHITECTURE.md §5):

      * Merged cells: only the anchor (top-left) cell keeps its value; the
        cells it "spans into" are left empty. This is the OPPOSITE of
        Docling's behaviour (which duplicates the merged value into every
        spanned position), and it avoids the "N×N inflation" that drives
        merged-cell-heavy sheets up to multi-MB outputs.
      * Empty-column pruning: 方眼紙 spec sheets often span 100+ columns
        with only a handful actually used. We collect the set of columns
        that hold real values and drop the rest entirely from the output —
        the table widens to match what's actually there, not the worksheet's
        nominal max_column.
      * Empty-row pruning: rows whose every (pruned) column is empty are
        skipped. Decorative blank rows between content blocks vanish.
      * Pipe / newline escaping: cell values are sanitized so Markdown
        parsing stays robust (newline → space; ``|`` → ``\\|``).
      * Embedded image anchors: when ``image_anchors`` is provided, a
        ``<!-- image: <filename> -->`` line is emitted AFTER the row that
        owns each image. This both feeds the Vision triage (which treats
        any ``<!-- image -->`` token as "send to Vision") AND keeps the
        filename next to the row that visually contains the image, so
        downstream RAG / chunk lineage can resolve picture references.

    Args:
      ws: openpyxl Worksheet.
      image_anchors: row(1-indexed) → list of asset filenames anchored at
        that row. Pass ``None`` (or empty dict) to disable image markers
        — preserves the original behaviour for callers that don't have
        anchor info (and keeps the function easy to unit-test).
      orphan_image_names: image filenames that exist in the xlsx zip but
        whose anchor cannot be resolved (e.g. EMF / WMF dropped by
        openpyxl). When non-empty, a footer block lists them so the
        information is not lost downstream — every chunk consumer can at
        least know the images exist and find them in ``assets/``.

    Returns a list of markdown lines (header + separator + body, with
    optional image marker lines and orphan footer interleaved) ready for
    joining with ``\\n``. Returns an empty list when the sheet has no
    content at all — caller decides whether to emit a stub or skip.
    """
    # Build a (row, col) → 'anchor' | 'spanned' map for merged regions.
    spanned: set[tuple[int, int]] = set()
    for mr in ws.merged_cells.ranges:
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                if not (r == mr.min_row and c == mr.min_col):
                    spanned.add((r, c))

    # First pass: harvest values, remember which columns are actually used.
    cell_values: dict[tuple[int, int], str] = {}
    cols_with_content: set[int] = set()
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0

    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            if (r, c) in spanned:
                continue
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            # Coerce to a single-line, pipe-safe string.
            text = str(v).replace("\n", " ").replace("\r", " ")
            text = text.replace("|", r"\|").strip()
            if text:
                cell_values[(r, c)] = text
                cols_with_content.add(c)

    # An image-only row (no cell text but an anchor sits on it) still
    # deserves a markdown line — without one, the image marker would
    # become orphaned and lose its "where in the sheet" context. So we
    # widen the set of "rows worth emitting" to include image anchor rows.
    image_anchors = image_anchors or {}
    anchor_rows = set(image_anchors.keys())

    if not cols_with_content and not anchor_rows:
        # Truly empty sheet — let caller emit a stub.
        # Even orphan images won't be useful without any context, so we
        # skip the footer in this corner case (caller's stub is enough).
        return []

    kept_cols = sorted(cols_with_content)

    # Second pass: emit rows in original row order. Track the source row
    # number for each emitted body row so we can interleave image markers.
    # When a row has no cell content but DOES anchor an image, we still
    # emit it as an empty row so the marker that follows lines up visually.
    emitted: list[tuple[int, list[str] | None]] = []   # (src_row, cells_or_None)
    for r in range(1, max_row + 1):
        row = [cell_values.get((r, c), "") for c in kept_cols] if kept_cols else []
        if any(row):
            emitted.append((r, row))
        elif r in anchor_rows:
            # Empty row but holds an image anchor → keep a placeholder row
            # so the marker has a home. ``None`` signals "no cell line".
            emitted.append((r, None))

    if not emitted:
        return []

    # Standard Markdown table: header row + separator + body rows.
    # We treat the first emitted row with cell content as the header.
    # If the very first emitted row is image-only (rare), we still emit
    # a 1-cell stub header so the rest of the table stays valid Markdown.
    n = max(len(kept_cols), 1)
    lines: list[str] = []

    # Helper: emit the image markers for a given source row, immediately
    # after the corresponding markdown row. Multiple images on the same
    # row each get their own line.
    #
    # Markers are wrapped as valid Markdown table rows
    # (``| <!-- image: ... --> |...``) rather than emitted as bare comment
    # lines. Why this matters: the SheetChunker's table-segment detector
    # treats any non-``|``-prefixed line as a hard segment boundary, so a
    # bare ``<!-- image: ... -->`` mid-table breaks the table in two and
    # makes the next data row look like a NEW header — which then gets
    # repeated at the top of every subsequent chunk in that sub-segment.
    # Real symptom seen on nra_kinou: row 137 ("５．２ 作業要員...") was
    # duplicated across chunks because an image marker right above it
    # split the table at that exact spot. Wrapping the marker in a table
    # row keeps the segment continuous; downstream consumers still see
    # the marker text intact (``<!-- image`` prefix), so chunk metadata's
    # ``has_image_ref`` and any RAG-side filename grep work unchanged.
    def _emit_markers(src_row: int) -> None:
        for fname in image_anchors.get(src_row, []):
            cells = [f"<!-- image: {fname} -->"] + [""] * (n - 1)
            lines.append("| " + " | ".join(cells) + " |")

    # Header (first emitted row).
    first_row_src, first_row_cells = emitted[0]
    if first_row_cells is not None:
        lines.append("| " + " | ".join(first_row_cells) + " |")
    else:
        # Image-only first row — emit a stub one-cell header so MD stays valid.
        lines.append("| |")
    lines.append("|" + "|".join(["---"] * n) + "|")
    _emit_markers(first_row_src)

    # Body.
    for src_row, cells in emitted[1:]:
        if cells is not None:
            lines.append("| " + " | ".join(cells) + " |")
        else:
            # Image-only body row — emit an empty table row so the marker
            # below it makes sense as "image at this position".
            lines.append("| " + " | ".join([""] * n) + " |")
        _emit_markers(src_row)

    # Orphan footer (only when we actually have orphans).
    if orphan_image_names:
        lines.append("")
        lines.append(
            "*Embedded images without resolvable anchor "
            "(see `assets/`):*"
        )
        for fname in orphan_image_names:
            # Keep the marker form so Vision/triage hooks still see them
            # as image references, but also list them as bullets for human
            # readers. One line each — tokens cost peanuts here.
            lines.append(f"- <!-- image: {fname} -->")

    return lines
